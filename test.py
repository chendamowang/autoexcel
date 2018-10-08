# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook('01.xlsx')
sheet1 = wb['2018']
sheet2 = wb['2017-5']
sheet3 = wb['2017-4']
wb2 = openpyxl.load_workbook('02.xlsx')
sheet0 = wb2.active

data = {}
for row in range(2, sheet1.max_row+1):
    ZJBH = sheet1['H'+str(row)].value
    DDH = sheet1['B'+str(row)].value
    SJBH = sheet1['C'+str(row)].value
    KHMC = sheet1['D'+str(row)].value
    KHJL = sheet1['E'+str(row)].value
    
    data.setdefault(ZJBH, {'订单号':None, '商机编号':None, '项目客户名称':None, '客户经理':None})
    data[ZJBH]['订单号'] = DDH
    data[ZJBH]['商机编号'] = SJBH
    data[ZJBH]['项目客户名称'] = KHMC
    data[ZJBH]['客户经理'] = KHJL

for row in range(2, sheet2.max_row+1):
    ZJBH = sheet2['H'+str(row)].value
    DDH = sheet2['B'+str(row)].value
    SJBH = sheet2['C'+str(row)].value
    KHMC = sheet2['D'+str(row)].value
    KHJL = sheet2['E'+str(row)].value
    
    data.setdefault(ZJBH, {'订单号':None, '商机编号':None, '项目客户名称':None, '客户经理':None})
    data[ZJBH]['订单号'] = DDH
    data[ZJBH]['商机编号'] = SJBH
    data[ZJBH]['项目客户名称'] = KHMC
    data[ZJBH]['客户经理'] = KHJL

for row in range(2, sheet3.max_row+1):
    ZJBH = sheet3['H'+str(row)].value
    DDH = sheet3['B'+str(row)].value
    SJBH = sheet3['C'+str(row)].value
    KHMC = sheet3['D'+str(row)].value
    KHJL = sheet3['E'+str(row)].value
    
    data.setdefault(ZJBH, {'订单号':None, '商机编号':None, '项目客户名称':None, '客户经理':None})
    data[ZJBH]['订单号'] = DDH
    data[ZJBH]['商机编号'] = SJBH
    data[ZJBH]['项目客户名称'] = KHMC
    data[ZJBH]['客户经理'] = KHJL

    
for row in range(3, sheet0.max_row+1):
    zjbh = sheet0['H'+str(row)].value
    if zjbh in data:
        sheet0['I'+str(row)].value = data[zjbh]['订单号']
        sheet0['J'+str(row)].value = data[zjbh]['商机编号']
        sheet0['K'+str(row)].value = data[zjbh]['项目客户名称']
        sheet0['L'+str(row)].value = data[zjbh]['客户经理']

wb2.save('02.xlsx')
